import urllib.request
from urllib import parse
import sys
from bs4 import BeautifulSoup
from selenium import webdriver
import os
import requests
import urllib.request as request
import time
import json
import argparse
import pickle
from pypinyin import lazy_pinyin
import re
import openpyxl
from tqdm import tqdm

def getPagebyName(name="Ruonan+Liu"):
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
    response = urllib.request.Request("https://dblp.uni-trier.de/search?q="+name, headers=headers)
    content = urllib.request.urlopen(response)
    # print(content.read().decode("utf-8"))
    soup = BeautifulSoup(content.read().decode("utf-8"), 'lxml')
    a = soup.find(text=name.split("+")[0]).parent.parent.parent.attrs['href']
    return a


def getPaperList(Personurl, name, update):


    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
    response = urllib.request.Request(Personurl, headers=headers)
    content = urllib.request.urlopen(response)
    # print(content.read().decode("utf-8"))
    soup = BeautifulSoup(content.read().decode("utf-8"), 'lxml')
    pub_list = soup.find_all(class_=re.compile("entry*"))
    papers = list()
    for paper in tqdm(pub_list):
        # 这里可以筛选年份 ，可以不适用目前
        year = paper.find(itemprop="datePublished").text
        title = paper.find(class_="title").text
        bibtex_url = paper.find('nav').find(text="BibTeX").parent.attrs['href']
        bibtex = getBibtexByUrl(bibtex_url)
        papers.append({'姓名':name,'论文名称':title,'BibTeX':bibtex,'年份':year})
    with open("data/"+name+".pickle", 'wb') as f:
        pickle.dump(papers, f)
    return papers


def getBibtexByUrl(url):
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
    response = urllib.request.Request(url, headers=headers)
    content = urllib.request.urlopen(response)
    # print(content.read().decode("utf-8"))
    soup = BeautifulSoup(content.read().decode("utf-8"), 'lxml')
    bibtex = soup.find('pre').text
    return bibtex


def printToExcel(filename, papers):
    f = openpyxl.Workbook()
    sheet = f['Sheet']
    col = [u'姓名', u'论文名称', u'BibTeX', u'年份']
    for jkey in tqdm(range(len(papers) + 1)):
        for i in range(len(col)):
            if jkey == 0:
                sheet.cell(jkey + 1, i + 1, value=col[i])

            else:
                sheet.cell(jkey + 1, i + 1, value=papers[jkey - 1][col[i]])
    f.save(filename)  # 保存文件


def namectonamee(namec):
    tmp = ""
    t = lazy_pinyin(namec)
    t[0] = t[0].capitalize()
    t[1] = t[1].capitalize()
    for i in range(1, len(t)):
        tmp += t[i]
    tmp += '+'
    tmp = tmp + t[0]
    return tmp


if __name__ == '__main__':
    # name = ['胡清华','韩亚洪','廖士中','邹权','谢宗霞','朱鹏飞','杨柳','王征','张长青','杨雅君','刘若楠','任冬伟','王旗龙','魏乐义','曹兵']
    parser = argparse.ArgumentParser()
    parser.add_argument('--m',type=int, help="what type? 1/2",default=1)
    parser.add_argument('--u', action="store_true",help='update?')
    parser.add_argument('--names', type=str, nargs='+')
    parser.add_argument('--person_url', type=str,nargs='+')
    args = parser.parse_args()
    names = args.names
    update = args.u
    m = args.m
    if m == 1:
        # 转化为英文
        for n in names:
            if not update:
                if os.path.exists("data/" + n + ".pickle"):
                    print("使用缓存！")
                    with open("data/" + n + ".pickle", 'rb') as f:
                        papers = pickle.load(f)
                else:
                    print("获取人物论文页：" + n)
                    personurl = getPagebyName(namectonamee(n))
                    print("成功:" + personurl + ".html")
                    print("开始获取论文列表:" + n)
                    papers = getPaperList(personurl + ".html", n, update=update)
            print("获取完毕！：")
            print("开始写入文件到"+n+".xlsx")
            printToExcel("results/"+n+".xlsx", papers)
            print("写入成功")
    # 直接指定人物页面
    elif m == 2:
        if args.person_url == "":
            print("请指定参数--person_url")
            exit()
        for i in range(len(args.person_url)):
            print("开始获取论文列表:" + names[i])
            papers = getPaperList(args.person_url, names[i], update=update)
            print("获取完毕！")
            print("开始写入文件到" + names[i] + ".xlsx")
            printToExcel("results/"+ names[i] + ".xlsx", papers)
            print("写入成功")






